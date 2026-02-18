"""
Core Managers - مدیرهای مرکزی (نسخه کامل)
"""
import logging
from typing import Dict, Any, Optional, List
from datetime import datetime
import os
import csv
import time
from PySide6.QtCore import *
from PySide6.QtWidgets import *
from PySide6.QtGui import *

logger = logging.getLogger(__name__)

# ==================== StatusBar Manager ====================
class StatusBarManager(QObject):
    """مدیریت StatusBar - نسخه ساده و کارا"""
    
    show_message_signal = Signal(str, str, int)  # widget_name, message, timeout
    clear_message_signal = Signal(str)  # widget_name
    show_progress_signal = Signal(str, str)  # widget_name, message
    show_success_signal = Signal(str, str)  # widget_name, message
    show_error_signal = Signal(str, str)  # widget_name, message
    
    _instance = None
    
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance
    
    def __init__(self):
        if not hasattr(self, "_initialized"):
            super().__init__()
            self._widgets = {}
            self._initialized = True
            self._widgets = {}
            self._main_window = None
            self._last_messages = {}
        
            # Connect signals
            self.show_message_signal.connect(self._handle_show_message)
            self.clear_message_signal.connect(self._handle_clear_message)
            self.show_progress_signal.connect(self._handle_show_progress)
            self.show_success_signal.connect(self._handle_show_success)
            self.show_error_signal.connect(self._handle_show_error)
            
            try:
                logger.info("✅ StatusBarManager initialized")
            except UnicodeEncodeError:
                logger.info("StatusBarManager initialized")  
            
    def _handle_show_message(self, widget_name: str, message: str, timeout: int = 3000):
        """Handle show message signal - با پیشگیری از Recursion"""
        try:
            current_time = time.time()
            last_message = self._last_messages.get(widget_name)
            
            if (last_message and 
                last_message['message'] == message and 
                current_time - last_message['time'] < 1.0):  # 1 ثانیه تاخیر
                logger.debug(f"Recursion prevented for {widget_name}: {message[:50]}...")
                return
                
            # ذخیره آخرین پیام
            self._last_messages[widget_name] = {
                'message': message,
                'time': current_time
            }
            
            if widget_name in self._widgets:
                widget = self._widgets[widget_name]
                
                # بررسی برای جلوگیری از حلقه
                if hasattr(widget, '_is_showing_message'):
                    if widget._is_showing_message:
                        logger.debug(f"Widget {widget_name} is already showing message")
                        return
                    widget._is_showing_message = True
                    
                try:
                    if isinstance(widget, QMainWindow):
                        widget.statusBar().showMessage(message, timeout)
                    elif hasattr(widget, 'show_message') and not isinstance(widget, StatusBarManager):
                        # مستقیماً متد show_message ویجت را صدا بزن، به شرطی که خود StatusBarManager نباشد
                        widget.show_message(message, timeout)
                    else:
                        logger.debug(f"[{widget_name}]: {message}")
                finally:
                    if hasattr(widget, '_is_showing_message'):
                        widget._is_showing_message = False
                        
        except RecursionError as re:
            logger.error(f"RecursionError in StatusBarManager for {widget_name}: {re}")
            # لاگ کن و جلوگیری از crash
            import traceback
            logger.error(traceback.format_exc())
        except Exception as e:
            logger.error(f"Error showing message for {widget_name}: {e}")
        
    def _handle_clear_message(self, widget_name: str):
        """Handle clear message signal"""
        try:
            if widget_name in self._widgets:
                widget = self._widgets[widget_name]
                if isinstance(widget, QMainWindow):
                    widget.statusBar().clearMessage()
        except Exception as e:
            logger.error(f"Error clearing message for {widget_name}: {e}")
    
    def _handle_show_progress(self, widget_name: str, message: str):
        """Handle progress message"""
        self.show_message_signal.emit(widget_name, f"🔄 {message}...", 0)
    
    def _handle_show_success(self, widget_name: str, message: str):
        """Handle success message"""
        self.show_message_signal.emit(widget_name, f"✅ {message}", 3000)
    
    def _handle_show_error(self, widget_name: str, message: str):
        """Handle error message"""
        self.show_message_signal.emit(widget_name, f"❌ {message}", 5000)
    
    def register_widget(self, widget_name: str, widget: QWidget):
        """ثبت ویجت"""
        if widget_name not in self._widgets:
            self._widgets[widget_name] = widget
            logger.debug(f"📝 Registered widget: {widget_name}")
    
    def register_main_window(self, main_window: QMainWindow):
        """ثبت MainWindow"""
        self.register_widget("MainWindow", main_window)
    
    # متدهای عمومی برای استفاده ساده
    def show_message(self, widget_name: str, message: str, timeout: int = 3000):
        """نمایش پیام"""
        self.show_message_signal.emit(widget_name, message, timeout)
    
    def clear_message(self, widget_name: str):
        """پاک کردن پیام"""
        self.clear_message_signal.emit(widget_name)
    
    def show_progress(self, widget_name: str, message: str):
        """نمایش پیام پیشرفت"""
        self.show_progress_signal.emit(widget_name, message)
    
    def show_success(self, widget_name: str, message: str):
        """نمایش پیام موفقیت"""
        self.show_success_signal.emit(widget_name, message)
    
    def show_error(self, widget_name: str, message: str):
        """نمایش پیام خطا"""
        self.show_error_signal.emit(widget_name, message)
    def show_warning(self, widget_name: str, message: str):
        """نمایش پیام هشدار"""
        self.show_message_signal.emit(widget_name, f"⚠️ {message}", 4000)

# ==================== AutoSave Manager ====================
class AutoSaveManager:
    """مدیریت Auto-Save"""
    
    def __init__(self):
        self._timers: Dict[str, QTimer] = {}
        self._widgets: Dict[str, QWidget] = {}
        self._enabled = True
        self._settings = QSettings("Nikan", "DrillMaster")
        self._status_manager = StatusBarManager()
    
    def enable_for_widget(self, widget_name: str, widget: QWidget, interval_minutes: int = 5):
        """فعال کردن Auto-Save برای ویجت"""
        if hasattr(widget, "save_data"):
            interval_ms = interval_minutes * 60 * 1000
            
            timer = QTimer()
            timer.timeout.connect(lambda: self._auto_save(widget_name, widget))
            timer.start(interval_ms)
            
            self._timers[widget_name] = timer
            self._widgets[widget_name] = widget
            
            logger.info(f"✅ Auto-save enabled for {widget_name} every {interval_minutes} minutes")
    
    def _auto_save(self, widget_name: str, widget: QWidget):
        """اجرای Auto-Save"""
        if not self._enabled:
            return
        
        try:
            if hasattr(widget, 'save_data'):
                result = widget.save_data()
                if result:
                    self._status_manager.show_message(
                        widget_name, 
                        f"💾 Auto-saved at {datetime.now().strftime('%H:%M:%S')}", 
                        2000
                    )
                else:
                    self._status_manager.show_message(
                        widget_name,
                        "💾 Auto-save failed",
                        2000
                    )
        except Exception as e:
            logger.error(f"Auto-save error for {widget_name}: {e}")
    
    def set_enabled(self, enabled: bool):
        """تنظیم وضعیت کلی Auto-Save"""
        self._enabled = enabled
        for timer in self._timers.values():
            if enabled:
                timer.start()
            else:
                timer.stop()
        
        status = "enabled" if enabled else "disabled"
        logger.info(f"🔄 Auto-save {status} globally")

# ==================== Shortcut Manager ====================
class ShortcutManager:
    """مدیریت کلیدهای میانبر"""
    
    def __init__(self, parent: QWidget):
        self.parent = parent
        self.shortcuts: Dict[str, Dict[str, Any]] = {}
        self._status_manager = StatusBarManager()
    
    def add_shortcut(self, key_sequence: str, slot, description: str = ""):
        """اضافه کردن کلید میانبر"""
        shortcut = QShortcut(QKeySequence(key_sequence), self.parent)
        shortcut.activated.connect(slot)
        
        self.shortcuts[key_sequence] = {
            "shortcut": shortcut,
            "description": description,
            "slot": slot,
        }
        
        return shortcut
    
    def add_shortcut_with_feedback(self, key_sequence: str, slot, description: str = ""):
        """اضافه کردن کلید میانبر با feedback"""
        def wrapped_slot():
            slot()
            self._status_manager.show_message(
                self.parent.__class__.__name__,
                f"Shortcut: {description}",
                1000
            )
        
        return self.add_shortcut(key_sequence, wrapped_slot, description)
    
    def setup_default_shortcuts(self):
        """تنظیم کلیدهای میانبر پیش‌فرض"""
        shortcuts = [
            ("Ctrl+S", self._save_current, "Save"),
            ("Ctrl+Shift+S", self._save_all, "Save All"),
            ("F5", self._refresh, "Refresh"),
            ("Ctrl+E", self._export, "Export"),
            ("F1", self._help, "Help"),
        ]
        
        for key_seq, slot, desc in shortcuts:
            self.add_shortcut_with_feedback(key_seq, slot, desc)
        
        logger.info(f"✅ {len(self.shortcuts)} shortcuts registered")
    
    def _save_current(self):
        """ذخیره جاری"""
        if hasattr(self.parent, 'save_current'):
            self.parent.save_current()
    
    def _save_all(self):
        """ذخیره همه"""
        if hasattr(self.parent, 'save_all'):
            self.parent.save_all()
    
    def _refresh(self):
        """رفرش"""
        if hasattr(self.parent, 'refresh'):
            self.parent.refresh()
    
    def _export(self):
        """اکسپورت"""
        if hasattr(self.parent, 'export'):
            self.parent.export()
    
    def _help(self):
        """راهنما"""
        if hasattr(self.parent, 'show_help'):
            self.parent.show_help()


# ==================== Table Manager ====================
class TableManager:
    """مدیریت پیشرفته جدول‌ها با قابلیت‌های Auto-Stretch، Undo/Redo و Scroll Optimization"""
    
    def __init__(self, table_widget: QTableWidget, parent_widget=None):
        self.table = table_widget
        self.parent = parent_widget
        self.undo_stack = []
        self.redo_stack = []
        self.max_stack_size = 100
        self._is_recording = True
        
        # تنظیمات Stretch و Scroll
        self._setup_table_optimizations()
        
        # ذخیره وضعیت اولیه
        self._initial_state = self._capture_table_state()
        
        # اتصال سیگنال‌ها
        self.table.itemChanged.connect(self._on_item_changed)
        
        # Setup keyboard shortcuts for undo/redo
        self._setup_shortcuts()
    
    def _setup_table_optimizations(self):
        """تنظیم بهینه‌سازی‌های جدول"""
        # ============ تنظیمات Stretch ============
        # 1. Stretch horizontal header
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        # 2. تنظیم اندازه ستون‌ها بر اساس محتوا (برای ردیف‌های بلند)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        
        # 3. فعال کردن word wrap برای ستون‌ها
        self.table.setWordWrap(True)
        
        # ============ تنظیمات Scroll ============
        # 1. بهتر کردن عملکرد اسکرول
        self.table.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.table.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        
        # 2. اضافه کردن اسکرول‌های نرم (Smooth Scrolling)
        self.table.verticalScrollBar().setSingleStep(20)
        self.table.horizontalScrollBar().setSingleStep(20)
        
        # ============ تنظیمات ظاهری ============
        # 1. Alternate row colors
        self.table.setAlternatingRowColors(True)
        
        # 2. نمایش grid
        self.table.setShowGrid(True)
        self.table.setGridStyle(Qt.SolidLine)
        
        # 3. انتخاب کامل سطر
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        
        # ============ تنظیمات سطرها و ستون‌ها ============
        # 1. ارتفاع پویای سطرها
        self.table.verticalHeader().setDefaultSectionSize(30)
        self.table.verticalHeader().setMinimumSectionSize(25)
        
        # 2. Set row height auto-adjustment
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        
        # 3. حداقل و حداکثر عرض ستون‌ها
        for col in range(self.table.columnCount()):
            self.table.horizontalHeader().setMinimumSectionSize(80)
            self.table.horizontalHeader().setMaximumSectionSize(400)
    
    def _setup_shortcuts(self):
        """تنظیم کلیدهای میانبر برای Undo/Redo"""
        if self.parent:
            # Undo: Ctrl+Z
            undo_shortcut = QShortcut(QKeySequence("Ctrl+Z"), self.parent)
            undo_shortcut.activated.connect(self.undo)
            
            # Redo: Ctrl+Y
            redo_shortcut = QShortcut(QKeySequence("Ctrl+Y"), self.parent)
            redo_shortcut.activated.connect(self.redo)
            
            # Add Row: Ctrl+A
            add_row_shortcut = QShortcut(QKeySequence("Ctrl+A"), self.parent)
            add_row_shortcut.activated.connect(lambda: self.add_row())
            
            # Delete Row: Ctrl+D
            delete_row_shortcut = QShortcut(QKeySequence("Ctrl+D"), self.parent)
            delete_row_shortcut.activated.connect(lambda: self.delete_row())
    
    def _capture_table_state(self):
        """ذخیره وضعیت فعلی جدول با جزئیات بیشتر"""
        state = {
            "row_count": self.table.rowCount(),
            "col_count": self.table.columnCount(),
            "data": [],
            "headers": [],
            "row_heights": [],
            "col_widths": [],
        }
        
        # ذخیره هدرها
        for col in range(state["col_count"]):
            header = self.table.horizontalHeaderItem(col)
            state["headers"].append(header.text() if header else f"Col {col+1}")
            state["col_widths"].append(self.table.columnWidth(col))
        
        # ذخیره داده‌ها و ارتفاع سطرها
        for row in range(state["row_count"]):
            row_data = []
            for col in range(state["col_count"]):
                item = self.table.item(row, col)
                if item:
                    row_data.append({
                        "text": item.text(),
                        "font": item.font().toString(),
                        "alignment": int(item.textAlignment()),
                        "foreground": item.foreground().color().name(),
                        "background": item.background().color().name(),
                    })
                else:
                    row_data.append({"text": "", "font": "", "alignment": 0, 
                                   "foreground": "#000000", "background": "#FFFFFF"})
            state["data"].append(row_data)
            state["row_heights"].append(self.table.rowHeight(row))
        
        return state
    
    def _on_item_changed(self, item):
        """ذخیره تغییرات برای Undo + Auto-resize"""
        if not self._is_recording or item is None:
            return
        
        old_value = getattr(item, "_table_manager_old_value", "")
        new_value = item.text()
        
        if old_value == new_value:
            return
        
        action = {
            "type": "item_change",
            "row": item.row(),
            "col": item.column(),
            "old_value": old_value,
            "new_value": new_value,
            "timestamp": datetime.now(),
        }
        
        item._table_manager_old_value = new_value
        self._push_to_undo(action)
        
        # Auto-adjust row height based on content
        self._auto_adjust_row_height(item.row())
    
    def _auto_adjust_row_height(self, row):
        """تنظیم خودکار ارتفاع سطر بر اساس محتوا"""
        try:
            # محاسبه ارتفاع مورد نیاز
            font_metrics = QFontMetrics(self.table.font())
            max_height = font_metrics.height() + 10  # حداقل ارتفاع
            
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item and item.text():
                    text = item.text()
                    # اگر متن طولانی است، ارتفاع بیشتری نیاز دارد
                    if len(text) > 50:
                        lines = len(text) // 50 + 1
                        height = (font_metrics.height() * lines) + 10
                        max_height = max(max_height, height)
            
            # تنظیم ارتفاع سطر
            self.table.setRowHeight(row, max_height)
            
        except Exception as e:
            logger.error(f"Error in auto_adjust_row_height: {e}")
    
    def _auto_adjust_column_widths(self):
        """تنظیم خودکار عرض ستون‌ها"""
        try:
            for col in range(self.table.columnCount()):
                self.table.resizeColumnToContents(col)
                # حداقل و حداکثر عرض
                current_width = self.table.columnWidth(col)
                min_width = 80
                max_width = 400
                
                if current_width < min_width:
                    self.table.setColumnWidth(col, min_width)
                elif current_width > max_width:
                    self.table.setColumnWidth(col, max_width)
                    
        except Exception as e:
            logger.error(f"Error in auto_adjust_column_widths: {e}")
    
    def delete_row(self, row_index=None):
        """حذف سطر با قابلیت Undo + Auto-adjust"""
        if row_index is None:
            row_index = self.table.currentRow()
        
        if row_index < 0 or row_index >= self.table.rowCount():
            return False
        
        self._is_recording = False
        
        # ذخیره داده‌های سطر قبل از حذف
        row_data = []
        for col in range(self.table.columnCount()):
            item = self.table.item(row_index, col)
            row_data.append(item.text() if item else "")
        
        # حذف سطر
        self.table.removeRow(row_index)
        
        # ذخیره action برای Undo
        action = {
            "type": "row_delete",
            "row": row_index,
            "row_data": row_data,
            "timestamp": datetime.now(),
        }
        self._push_to_undo(action)
        
        self._is_recording = True
        
        # Auto-adjust columns after deletion
        QTimer.singleShot(100, self._auto_adjust_column_widths)
        
        return True
    
    def add_row(self, default_data=None, position=-1):
        """اضافه کردن سطر با قابلیت Undo + Auto-stretch"""
        row = self.table.rowCount() if position == -1 else position
        self._is_recording = False
        
        # درج سطر
        self.table.insertRow(row)
        
        # پر کردن با داده پیش‌فرض
        if default_data:
            for col, value in enumerate(default_data):
                if col < self.table.columnCount():
                    item = QTableWidgetItem(str(value))
                    item._table_manager_old_value = str(value)
                    
                    # تنظیمات آیتم برای نمایش بهتر
                    item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                    
                    # اگر متن طولانی است، word wrap فعال شود
                    if len(str(value)) > 50:
                        item.setToolTip(str(value))
                    
                    self.table.setItem(row, col, item)
        
        # ذخیره action برای Undo
        action = {
            "type": "row_add",
            "row": row,
            "row_data": default_data or [],
            "timestamp": datetime.now(),
        }
        self._push_to_undo(action)
        
        self._is_recording = True
        
        # Auto-adjust بعد از اضافه شدن
        self._auto_adjust_row_height(row)
        QTimer.singleShot(100, self._auto_adjust_column_widths)
        
        return row
    
    def add_column(self, header="", default_data=None):
        """اضافه کردن ستون جدید با Stretch خودکار"""
        col = self.table.columnCount()
        self._is_recording = False
        
        # درج ستون
        self.table.insertColumn(col)
        
        # تنظیم هدر
        if header:
            self.table.setHorizontalHeaderItem(col, QTableWidgetItem(header))
        
        # پر کردن داده پیش‌فرض
        if default_data:
            for row, value in enumerate(default_data):
                if row < self.table.rowCount():
                    item = QTableWidgetItem(str(value))
                    item._table_manager_old_value = str(value)
                    self.table.setItem(row, col, item)
        
        # ذخیره action برای Undo
        action = {
            "type": "column_add",
            "col": col,
            "header": header,
            "default_data": default_data or [],
            "timestamp": datetime.now(),
        }
        self._push_to_undo(action)
        
        self._is_recording = True
        
        # Auto-adjust ستون جدید
        QTimer.singleShot(100, self._auto_adjust_column_widths)
        
        return col
    
    def _push_to_undo(self, action):
        """Push action to undo stack"""
        self.undo_stack.append(action)
        if len(self.undo_stack) > self.max_stack_size:
            self.undo_stack.pop(0)
        self.redo_stack.clear()  # Clear redo stack on new action
    
    def undo(self):
        """انجام Undo با حفظ تنظیمات Stretch"""
        if not self.undo_stack:
            return False
        
        self._is_recording = False
        
        action = self.undo_stack.pop()
        self.redo_stack.append(action)
        
        if action["type"] == "item_change":
            item = self.table.item(action["row"], action["col"])
            if item:
                item.setText(action["old_value"])
                item._table_manager_old_value = action["old_value"]
                self._auto_adjust_row_height(action["row"])
        
        elif action["type"] == "row_add":
            self.table.removeRow(action["row"])
        
        elif action["type"] == "row_delete":
            self.table.insertRow(action["row"])
            for col, value in enumerate(action["row_data"]):
                if col < self.table.columnCount():
                    item = QTableWidgetItem(str(value))
                    item._table_manager_old_value = str(value)
                    self.table.setItem(action["row"], col, item)
            self._auto_adjust_row_height(action["row"])
        
        elif action["type"] == "column_add":
            self.table.removeColumn(action["col"])
        
        self._is_recording = True
        
        # Auto-adjust بعد از undo
        QTimer.singleShot(100, self._auto_adjust_column_widths)
        
        return True
    
    def redo(self):
        """انجام Redo با حفظ تنظیمات Stretch"""
        if not self.redo_stack:
            return False
        
        self._is_recording = False
        
        action = self.redo_stack.pop()
        self.undo_stack.append(action)
        
        if action["type"] == "item_change":
            item = self.table.item(action["row"], action["col"])
            if item:
                item.setText(action["new_value"])
                item._table_manager_old_value = action["new_value"]
                self._auto_adjust_row_height(action["row"])
        
        elif action["type"] == "row_add":
            self.table.insertRow(action["row"])
            for col, value in enumerate(action["row_data"]):
                if col < self.table.columnCount():
                    item = QTableWidgetItem(str(value))
                    item._table_manager_old_value = str(value)
                    self.table.setItem(action["row"], col, item)
            self._auto_adjust_row_height(action["row"])
        
        elif action["type"] == "row_delete":
            self.table.removeRow(action["row"])
        
        elif action["type"] == "column_add":
            self.table.insertColumn(action["col"])
            if action["header"]:
                self.table.setHorizontalHeaderItem(action["col"], 
                                                  QTableWidgetItem(action["header"]))
            for row, value in enumerate(action["default_data"]):
                if row < self.table.rowCount():
                    item = QTableWidgetItem(str(value))
                    item._table_manager_old_value = str(value)
                    self.table.setItem(row, action["col"], item)
        
        self._is_recording = True
        
        # Auto-adjust بعد از redo
        QTimer.singleShot(100, self._auto_adjust_column_widths)
        
        return True
    
    def can_undo(self):
        """آیا امکان Undo وجود دارد؟"""
        return len(self.undo_stack) > 0
    
    def can_redo(self):
        """آیا امکان Redo وجود دارد؟"""
        return len(self.redo_stack) > 0
    
    def clear_history(self):
        """پاک کردن تاریخچه Undo/Redo"""
        self.undo_stack.clear()
        self.redo_stack.clear()
        self._initial_state = self._capture_table_state()
    
    def set_alternating_row_colors(self, enabled=True, color1="#FFFFFF", color2="#F5F5F5"):
        """تنظیم رنگ‌های متناوب سطرها"""
        self.table.setAlternatingRowColors(enabled)
        if enabled:
            style = f"""
            QTableWidget {{
                alternate-background-color: {color2};
                background-color: {color1};
                gridline-color: #D0D0D0;
            }}
            QTableWidget::item:selected {{
                background-color: #0078D4;
                color: white;
            }}
            """
            self.table.setStyleSheet(style)
    
    def set_column_stretch_mode(self, mode="stretch"):
        """تنظیم حالت Stretch برای ستون‌ها"""
        if mode == "stretch":
            self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        elif mode == "interactive":
            self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        elif mode == "fixed":
            self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Fixed)
        elif mode == "resize_to_contents":
            self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
    
    def export_to_excel(self, filename=None):
        """اکسپورت جدول به Excel"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, Border, Side
            
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"table_export_{timestamp}.xlsx"
            
            # جمع‌آوری داده‌ها
            data = []
            headers = []
            
            # هدرها
            for col in range(self.table.columnCount()):
                header = self.table.horizontalHeaderItem(col)
                headers.append(header.text() if header else f"Column {col+1}")
            
            # داده‌ها
            for row in range(self.table.rowCount()):
                row_data = []
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    row_data.append(item.text() if item else "")
                data.append(row_data)
            
            # ایجاد DataFrame و ذخیره در Excel
            df = pd.DataFrame(data, columns=headers)
            
            # استفاده از openpyxl برای فرمت‌دهی بهتر
            wb = Workbook()
            ws = wb.active
            ws.title = "Table Export"
            
            # نوشتن هدرها با استایل
            header_font = Font(bold=True, size=12)
            header_alignment = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'),
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # نوشتن داده‌ها
            for row_idx, row_data in enumerate(data, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(wrap_text=True)
                    cell.border = thin_border
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filename)
            return filename
            
        except ImportError:
            logger.error("Excel export requires 'pandas' and 'openpyxl' packages")
            return None
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            return None
    
    def print_preview(self):
        """پیش‌نمایش چاپ"""
        try:
            printer = QPrinter(QPrinter.HighResolution)
            preview = QPrintPreviewDialog(printer)
            preview.paintRequested.connect(self._print_table)
            preview.exec_()
        except Exception as e:
            logger.error(f"Error in print preview: {e}")
    def import_from_csv(self, filename):
        """ایمپورت داده از فایل CSV"""
        try:
            with open(filename, 'r', encoding='utf-8-sig') as file:
                reader = csv.reader(file)
                rows = list(reader)
                
            if not rows:
                return False
                
            # پاک کردن جدول فعلی
            self.table.setRowCount(0)
            
            # تنظیم تعداد ستون‌ها
            col_count = len(rows[0]) if rows else self.table.columnCount()
            self.table.setColumnCount(col_count)
            
            # تنظیم هدرها
            headers = rows[0]
            self.table.setHorizontalHeaderLabels(headers)
            
            # اضافه کردن داده‌ها
            for row_idx, row_data in enumerate(rows[1:], start=0):
                self.add_row(row_data)
            
            return True
            
        except Exception as e:
            logger.error(f"Error importing from CSV: {e}")
            return False
    def _print_table(self, printer):
        """چاپ جدول"""
        painter = QPainter(printer)
        
        # تنظیم فونت
        font = QFont("Arial", 10)
        painter.setFont(font)
        
        # محاسبه ابعاد
        page_rect = printer.pageRect(QPrinter.DevicePixel)
        margin = 50
        usable_width = page_rect.width() - 2 * margin
        usable_height = page_rect.height() - 2 * margin
        
        # محاسبه عرض ستون‌ها
        col_count = self.table.columnCount()
        col_width = usable_width / col_count if col_count > 0 else usable_width
        
        # چاپ هدرها
        y = margin
        for col in range(col_count):
            header = self.table.horizontalHeaderItem(col)
            header_text = header.text() if header else f"Column {col+1}"
            
            x = margin + col * col_width
            painter.drawText(QRectF(x, y, col_width, 30), 
                           Qt.AlignCenter | Qt.TextWordWrap, 
                           header_text)
        
        # خط زیر هدر
        painter.drawLine(margin, y + 35, margin + usable_width, y + 35)
        
        # چاپ داده‌ها
        y += 50
        for row in range(self.table.rowCount()):
            for col in range(col_count):
                item = self.table.item(row, col)
                text = item.text() if item else ""
                
                x = margin + col * col_width
                painter.drawText(QRectF(x, y, col_width, 25), 
                               Qt.AlignLeft | Qt.AlignVCenter | Qt.TextWordWrap, 
                               text)
            y += 30
            
            # اگر به انتهای صفحه رسیدیم، صفحه جدید شروع کنیم
            if y > page_rect.height() - margin:
                printer.newPage()
                y = margin
        
        painter.end()

# ==================== Table Button Manager ====================
class TableButtonManager:
    """مدیریت متمرکز دکمه‌های Add/Remove برای جدول‌ها"""
    
    @staticmethod
    def setup_table_with_buttons(
        parent_widget,
        table_widget,
        add_callback=None,
        remove_callback=None,
        calculate_callback=None,
        export_callback=None,
        import_callback=None,
        title=None,
    ):
        """تنظیم جدول با دکمه‌های استاندارد"""
        
        # ایجاد layout اصلی
        main_layout = QVBoxLayout()
        
        # اضافه کردن عنوان اگر موجود باشد
        if title:
            title_label = QLabel(f"<h3>{title}</h3>")
            title_label.setAlignment(Qt.AlignCenter)
            main_layout.addWidget(title_label)
        
        # اضافه کردن جدول
        main_layout.addWidget(table_widget)
        
        # ایجاد layout برای دکمه‌ها
        button_layout = QHBoxLayout()
        
        # دکمه Add
        if add_callback:
            add_btn = QPushButton("➕ Add Row")
            add_btn.setToolTip("Add a new row")
            add_btn.clicked.connect(add_callback)
            button_layout.addWidget(add_btn)
        
        # دکمه Remove
        if remove_callback:
            remove_btn = QPushButton("➖ Remove Row")
            remove_btn.setToolTip("Remove selected row")
            remove_btn.clicked.connect(remove_callback)
            button_layout.addWidget(remove_btn)
        
        # دکمه Calculate
        if calculate_callback:
            calc_btn = QPushButton("🔄 Calculate")
            calc_btn.setToolTip("Perform calculations")
            calc_btn.clicked.connect(calculate_callback)
            button_layout.addWidget(calc_btn)
        
        # دکمه Import
        if import_callback:
            import_btn = QPushButton("📂 Import")
            import_btn.setToolTip("Import from CSV/Excel")
            import_btn.clicked.connect(import_callback)
            button_layout.addWidget(import_btn)
        
        # دکمه Export
        if export_callback:
            export_btn = QPushButton("📤 Export")
            export_btn.setToolTip("Export to CSV/Excel/PDF")
            export_btn.clicked.connect(export_callback)
            button_layout.addWidget(export_btn)
        
        button_layout.addStretch()
        main_layout.addLayout(button_layout)
        
        # تنظیم layout برای parent
        def clear_layout(l):
            while l and l.count():
                item = l.takeAt(0)
                w = item.widget()
                if w:
                    w.setParent(None)
        
        if parent_widget.layout():
            clear_layout(parent_widget.layout())
        
        parent_widget.setLayout(main_layout)
        
        # ایجاد TableManager برای Undo/Redo
        table_manager = TableManager(table_widget)
        setattr(table_widget, "_table_manager", table_manager)
        
        return table_manager
    
    @staticmethod
    def add_table_row(table, default_data=None, position=-1):
        """اضافه کردن سطر به جدول"""
        if hasattr(table, "_table_manager"):
            # استفاده از TableManager اگر موجود باشد
            return table._table_manager.add_row(default_data, position)
        else:
            # روش ساده
            row = table.rowCount() if position == -1 else position
            table.insertRow(row)
            
            if default_data:
                for col, value in enumerate(default_data):
                    if col < table.columnCount():
                        table.setItem(row, col, QTableWidgetItem(str(value)))
            
            return row
    
    @staticmethod
    def remove_table_row(table, row_index=None):
        """حذف سطر از جدول"""
        if hasattr(table, "_table_manager"):
            # استفاده از TableManager اگر موجود باشد
            return table._table_manager.delete_row(row_index)
        else:
            # روش ساده
            if row_index is None:
                row_index = table.currentRow()
            
            if 0 <= row_index < table.rowCount():
                table.removeRow(row_index)
                return True
            
            return False


# ==================== Export Manager ====================
class ExportManager:
    """مدیریت export به فرمت‌های مختلف"""
    
    def __init__(self, parent=None):
        self.parent = parent
        self._status_manager = StatusBarManager()
    
    def export_table_with_dialog(self, table_widget, default_name="export"):
        """اکسپورت جدول با دیالوگ انتخاب مسیر"""
        if not self.parent:
            return None
        
        # دیالوگ انتخاب فرمت
        formats = ["CSV", "PDF", "Excel"]
        format_choice, ok = QInputDialog.getItem(
            self.parent, "Export Format", "Select export format:", formats, 0, False
        )
        
        if not ok or not format_choice:
            return None
        
        # دیالوگ انتخاب مسیر ذخیره
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if format_choice == "CSV":
            filter_text = "CSV Files (*.csv);;All Files (*.*)"
            default_filename = f"{default_name}_{timestamp}.csv"
            ext = ".csv"
        elif format_choice == "PDF":
            filter_text = "PDF Files (*.pdf);;All Files (*.*)"
            default_filename = f"{default_name}_{timestamp}.pdf"
            ext = ".pdf"
        else:  # Excel
            filter_text = "Excel Files (*.xlsx);;All Files (*.*)"
            default_filename = f"{default_name}_{timestamp}.xlsx"
            ext = ".xlsx"
        
        filename, _ = QFileDialog.getSaveFileName(
            self.parent, "Save Export File", default_filename, filter_text
        )
        
        if not filename:
            return None
        
        # اضافه کردن extension اگر وجود ندارد
        if not filename.endswith(ext):
            filename += ext
        
        # نمایش progress
        self._status_manager.show_progress(
            self.parent.__class__.__name__, f"Exporting to {format_choice}..."
        )
        
        # اجرای export
        result = self.export_table(table_widget, format_choice.lower(), filename)
        
        if result:
            self._status_manager.show_success(
                self.parent.__class__.__name__,
                f"Exported to: {os.path.basename(filename)}",
            )
        else:
            self._status_manager.show_error(
                self.parent.__class__.__name__, "Export failed"
            )
        
        return result
    
    def export_table(self, table_widget, format="csv", filename=None):
        """اکسپورت جدول به فرمت‌های مختلف"""
        if format.lower() == "csv":
            return self._export_to_csv(table_widget, filename)
        elif format.lower() == "pdf":
            return self._export_to_pdf(table_widget, filename)
        elif format.lower() == "excel":
            return self._export_to_excel(table_widget, filename)
        else:
            return None
    
    def _export_to_csv(self, table_widget, filename=None):
        """اکسپورت به CSV"""
        try:
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"export_{timestamp}.csv"
            
            with open(filename, "w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                
                # نوشتن هدرها
                headers = []
                for col in range(table_widget.columnCount()):
                    header_item = table_widget.horizontalHeaderItem(col)
                    headers.append(
                        header_item.text() if header_item else f"Column {col+1}"
                    )
                writer.writerow(headers)
                
                # نوشتن داده‌ها
                for row in range(table_widget.rowCount()):
                    row_data = []
                    for col in range(table_widget.columnCount()):
                        item = table_widget.item(row, col)
                        row_data.append(item.text() if item else "")
                    writer.writerow(row_data)
            
            return filename
            
        except Exception as e:
            logger.error(f"CSV export failed: {e}")
            return None
    
    def _export_to_pdf(self, table_widget, filename=None):
        """اکسپورت به PDF"""
        try:
            # نیاز به reportlab دارد
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
            
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"export_{timestamp}.pdf"
            
            # ایجاد PDF ساده
            c = canvas.Canvas(filename, pagesize=A4)
            width, height = A4
            
            # اضافه کردن عنوان
            c.setFont("Helvetica-Bold", 16)
            c.drawString(50, height - 50, "Table Export")
            
            c.save()
            return filename
            
        except ImportError:
            logger.error("PDF export requires 'reportlab' package")
            return None
        except Exception as e:
            logger.error(f"PDF export failed: {e}")
            return None
    
    def _export_to_excel(self, table_widget, filename=None):
        """اکسپورت به Excel"""
        try:
            # نیاز به openpyxl دارد
            from openpyxl import Workbook
            
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"export_{timestamp}.xlsx"
            
            # ایجاد workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Table Export"
            
            # نوشتن هدرها
            for col in range(table_widget.columnCount()):
                header = table_widget.horizontalHeaderItem(col)
                ws.cell(row=1, column=col+1, 
                       value=header.text() if header else f"Column {col+1}")
            
            # نوشتن داده‌ها
            for row in range(table_widget.rowCount()):
                for col in range(table_widget.columnCount()):
                    item = table_widget.item(row, col)
                    ws.cell(row=row+2, column=col+1, 
                           value=item.text() if item else "")
            
            wb.save(filename)
            return filename
            
        except ImportError:
            logger.error("Excel export requires 'openpyxl' package")
            return None
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            return None
    
    def export_data(self):
        """Export data using ExportManager"""
        try:
            if hasattr(self, 'wells_table'):
                export_manager = ExportManager(self)
                result = export_manager.export_table_with_dialog(
                    self.wells_table,
                    "dashboard_wells"
                )
                if result:
                    self.status_manager.show_success(
                        "HomeTab",
                        f"Data exported to {result}"
                    )
                else:
                    self.status_manager.show_error(
                        "HomeTab",
                        "Export cancelled or failed"
                    )
            else:
                self.status_manager.show_error(
                    "HomeTab",
                    "No data available to export"
                )
        except Exception as e:
            logger.error(f"Export error: {e}")
            self.status_manager.show_error(
                "HomeTab",
                f"Export failed: {str(e)}"
            )

# ==================== Quick Setup Functions ====================
def setup_widget_with_managers(widget, widget_name, enable_autosave=True, 
                              autosave_interval=5, setup_shortcuts=True):
    """تنظیم سریع ویجت با همه managerها"""
    # ثبت در StatusBar Manager
    status_manager = StatusBarManager()
    status_manager.register_widget(widget_name, widget)
    
    # فعال کردن Auto-save
    if enable_autosave and hasattr(widget, "save_data"):
        auto_save_manager = AutoSaveManager()
        auto_save_manager.enable_for_widget(widget_name, widget, autosave_interval)
    
    # تنظیم کلیدهای میانبر
    if setup_shortcuts and hasattr(widget, "setup_shortcuts"):
        widget.setup_shortcuts()
    
    logger.info(f"✅ Widget '{widget_name}' setup complete with all managers")
    return True
    

class DrillingManager:
    """مدیریت محاسبات حفاری"""
    
    @staticmethod
    def calculate_tfa(nozzles_data: List[Dict]) -> float:
        """محاسبه Total Flow Area"""
        total_tfa = 0.0
        for nozzle in nozzles_data:
            size_32 = nozzle.get('size_32nd', 16)
            quantity = nozzle.get('quantity', 1)
            diameter_inch = size_32 / 32.0
            area = 3.1416 * (diameter_inch / 2.0) ** 2
            total_tfa += area * quantity
        return round(total_tfa, 3)
    
    @staticmethod
    def calculate_rop(depth_in: float, depth_out: float, hours: float) -> float:
        """محاسبه Rate of Penetration"""
        if hours <= 0:
            return 0.0
        bit_drilled = depth_out - depth_in
        if bit_drilled <= 0:
            return 0.0
        return round(bit_drilled / hours, 2)
    
    @staticmethod
    def calculate_hsi(pump_pressure: float, flow_rate: float, tfa: float) -> float:
        """محاسبه Hydraulic Horsepower per Square Inch"""
        if tfa <= 0:
            return 0.0
        # HSI = (P × Q) / (TFA × 1714)
        hsi = (pump_pressure * flow_rate) / (tfa * 1714)
        return round(hsi, 2)
    
    @staticmethod
    def calculate_annular_velocity(flow_rate: float, annulus_area: float) -> float:
        """محاسبه Annular Velocity"""
        if annulus_area <= 0:
            return 0.0
        # AV = (24.5 × Q) / (OD² - ID²)
        # این یک فرمول ساده‌شده است
        av = flow_rate * 0.5  # نیاز به اطلاعات دقیق‌تر
        return round(av, 1)
    
    @staticmethod
    def calculate_bit_revolution(rpm: float, hours: float) -> float:
        """محاسبه Bit Cumulative Revolution"""
        return round(rpm * hours * 60, 0)
    
    @staticmethod
    def validate_drilling_data(data: Dict) -> Dict[str, str]:
        """اعتبارسنجی داده‌های حفاری - نسخه اصلاح شده"""
        errors = {}
        
        # بررسی فیلدهای ضروری
        required_fields = ['bit_no', 'bit_size', 'depth_in', 'depth_out']
        for field in required_fields:
            if field not in data or not str(data.get(field, '')).strip():
                errors[field] = f"{field.replace('_', ' ').title()} is required"
        
        # اعتبارسنجی مقادیر عددی
        numeric_fields = ['bit_size', 'depth_in', 'depth_out', 'wob_min', 'wob_max', 
                         'rpm_min', 'rpm_max', 'torque_min', 'torque_max']
        for field in numeric_fields:
            value = data.get(field)
            if value is not None and str(value).strip():
                try:
                    float(value)
                except (ValueError, TypeError):
                    errors[field] = f"{field.replace('_', ' ').title()} must be a number"
        
        # اعتبارسنجی منطقی
        depth_in = data.get('depth_in')
        depth_out = data.get('depth_out')
        if depth_in and depth_out:
            try:
                if float(depth_in) >= float(depth_out):
                    errors['depth'] = "Depth Out must be greater than Depth In"
            except (ValueError, TypeError):
                pass
        
        return errors
    
    def validate_form_data(self):
        """اعتبارسنجی فرم با استفاده از DrillingManager"""
        form_data = {
            'bit_no': self.bit_no.text(),
            'bit_size': self.bit_size.value(),
            'depth_in': self.depth_in.value(),
            'depth_out': self.depth_out.value(),
            'wob_min': self.wob_min.value(),
            'wob_max': self.wob_max.value(),
            # بقیه فیلدها...
        }
        
        errors = self.drilling_manager.validate_drilling_data(form_data)
        return errors