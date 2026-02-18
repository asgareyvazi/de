"""
Equipment Widget - ویجت تجهیزات با قابلیت‌های کامل
هر تب به صورت کلاس جداگانه
"""

import logging
import json
from datetime import datetime, date
from typing import Dict, Any, List, Optional

from PySide6.QtWidgets import *
from PySide6.QtCore import *
from PySide6.QtGui import *

from core.managers import StatusBarManager, TableManager, ExportManager, TableButtonManager
from core.database import DailyReport, Well

logger = logging.getLogger(__name__)


# ------------------------ Base Equipment Widget ------------------------
class DrillWidgetBase(QWidget):
    """Base class for all drill widgets with common functionality"""
    
    def __init__(self, widget_name, db_manager=None):
        super().__init__()
        self.widget_name = widget_name
        self.db_manager = db_manager
        self.status_manager = StatusBarManager()
        self.status_manager.register_widget(widget_name, self)
        
    def save_data(self):
        """Base save method - should be overridden"""
        return True
        
    def load_data(self):
        """Base load method - should be overridden"""
        return True
        
    def setup_shortcuts(self):
        """Setup keyboard shortcuts"""
        pass


# ------------------------ Rig Equipment Tab ------------------------
class RigEquipmentTab(QWidget):
    """تب تجهیزات ریگ"""
    
    def __init__(self, parent_widget=None, db_manager=None):
        super().__init__()
        self.parent_widget = parent_widget
        self.db_manager = db_manager
        self.table_manager = None
        self.init_ui()
        
    def init_ui(self):
        """راه‌اندازی رابط کاربری"""
        layout = QVBoxLayout()
        
        # توضیحات
        desc_label = QLabel("🏗️ Rig Equipment - Track all rig equipment with details and maintenance history")
        desc_label.setStyleSheet("font-size: 12px; color: #555; padding: 5px;")
        layout.addWidget(desc_label)
        
        # جدول تجهیزات ریگ
        self.table = QTableWidget(0, 8)
        self.table.setHorizontalHeaderLabels([
            "Equipment", "Category", "Model", "Serial No", 
            "Status", "Last Maintenance", "Next Maintenance", "Notes"
        ])
        
        # تنظیم Table Manager
        self.table_manager = TableManager(self.table, self)
        
        # تنظیم دکمه‌ها
        button_widget = self.create_buttons()
        
        layout.addWidget(self.table)
        layout.addWidget(button_widget)
        
        self.setLayout(layout)
        
        # اضافه کردن نمونه داده
        self.load_sample_data()
        
    def create_buttons(self):
        """ایجاد دکمه‌های تب"""
        widget = QWidget()
        layout = QHBoxLayout()
        
        add_btn = QPushButton("➕ Add Equipment")
        add_btn.clicked.connect(self.add_row)
        
        remove_btn = QPushButton("➖ Remove Equipment")
        remove_btn.clicked.connect(self.remove_row)
        
        save_btn = QPushButton("💾 Save")
        save_btn.clicked.connect(self.save_data)
        
        load_btn = QPushButton("📥 Load")
        load_btn.clicked.connect(self.load_data)
        
        export_btn = QPushButton("📤 Export")
        export_btn.clicked.connect(self.export_data)
        
        layout.addWidget(add_btn)
        layout.addWidget(remove_btn)
        layout.addStretch()
        layout.addWidget(save_btn)
        layout.addWidget(load_btn)
        layout.addWidget(export_btn)
        
        widget.setLayout(layout)
        return widget
        
    def load_sample_data(self):
        """بارگذاری داده نمونه"""
        if self.table.rowCount() == 0:
            sample_data = [
                ["Drawworks", "Hoisting", "DW-1000", "SN001", 
                 "Operational", "2024-01-15", "2024-04-15", "Regular maintenance completed"],
                ["Mud Pump", "Pumping", "MP-800", "SN002", 
                 "Operational", "2024-02-01", "2024-05-01", "Replaced seals"],
                ["Top Drive", "Rotary", "TD-750", "SN003", 
                 "Operational", "2024-01-30", "2024-04-30", "Calibrated"],
                ["BOP", "Safety", "BOP-13-5/8", "SN004", 
                 "Operational", "2024-03-01", "2024-06-01", "Function test passed"],
                ["Crane", "Lifting", "C-50T", "SN005", 
                 "Maintenance", "2024-02-15", "2024-03-15", "Scheduled maintenance"]
            ]
            
            for data in sample_data:
                self.add_row(data)
                
    def add_row(self, data=None):
        """اضافه کردن سطر"""
        if data is None:
            data = ["New Equipment", "Category", "Model", 
                   f"SN{self.table.rowCount()+1:03d}", 
                   "Operational", datetime.now().strftime("%Y-%m-%d"), 
                   "", ""]
        
        self.table_manager.add_row(data)
        
    def remove_row(self):
        """حذف سطر"""
        self.table_manager.delete_row()
        
    def save_data(self):
        """ذخیره داده‌ها"""
        try:
            data = self.get_table_data()
            # ذخیره در parent widget یا دیتابیس
            if self.parent_widget and hasattr(self.parent_widget, 'save_rig_equipment'):
                return self.parent_widget.save_rig_equipment(data)
                
            QMessageBox.information(self, "Success", "Rig equipment data saved")
            return True
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to save: {str(e)}")
            return False
            
    def load_data(self):
        """بارگذاری داده‌ها"""
        try:
            # بارگذاری از parent widget یا دیتابیس
            if self.parent_widget and hasattr(self.parent_widget, 'load_rig_equipment'):
                data = self.parent_widget.load_rig_equipment()
                if data:
                    self.load_table_data(data)
                    return True
                    
            QMessageBox.information(self, "Info", "No data to load")
            return False
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to load: {str(e)}")
            return False
            
    def export_data(self):
        """اکسپورت داده‌ها"""
        export_manager = ExportManager(self)
        result = export_manager.export_table_with_dialog(self.table, "rig_equipment")
        if result:
            QMessageBox.information(self, "Success", f"Exported to {result}")
            
    def get_table_data(self):
        """دریافت داده‌های جدول"""
        data = []
        for row in range(self.table.rowCount()):
            row_data = []
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                row_data.append(item.text() if item else "")
            data.append(row_data)
        return data
        
    def load_table_data(self, data):
        """بارگذاری داده‌ها در جدول"""
        self.table.setRowCount(0)
        for row_data in data:
            row = self.table.rowCount()
            self.table.insertRow(row)
            for col, value in enumerate(row_data):
                if col < self.table.columnCount():
                    item = QTableWidgetItem(str(value))
                    self.table.setItem(row, col, item)


# ------------------------ Inventory Tab ------------------------
class InventoryTab(QWidget):
    """تب موجودی"""
    
    def __init__(self, parent_widget=None, db_manager=None):
        super().__init__()
        self.parent_widget = parent_widget
        self.db_manager = db_manager
        self.table_manager = None
        self.init_ui()
        
    def init_ui(self):
        """راه‌اندازی رابط کاربری"""
        layout = QVBoxLayout()
        
        # توضیحات
        desc_label = QLabel("📦 Inventory Management - Track materials, chemicals, and supplies with stock levels")
        desc_label.setStyleSheet("font-size: 12px; color: #555; padding: 5px;")
        layout.addWidget(desc_label)
        
        # جدول موجودی
        self.table = QTableWidget(0, 9)
        self.table.setHorizontalHeaderLabels([
            "Item", "Category", "Opening Stock", "Received", 
            "Used", "Remaining", "Unit", "Min Level", "Max Level"
        ])
        
        # Table Manager
        self.table_manager = TableManager(self.table, self)
        
        # دکمه‌ها
        button_widget = self.create_buttons()
        
        layout.addWidget(self.table)
        layout.addWidget(button_widget)
        
        self.setLayout(layout)
        
        # نمونه داده
        self.load_sample_data()
        
    def create_buttons(self):
        """ایجاد دکمه‌های تب"""
        widget = QWidget()
        layout = QHBoxLayout()
        
        add_btn = QPushButton("➕ Add Item")
        add_btn.clicked.connect(self.add_row)
        
        remove_btn = QPushButton("➖ Remove Item")
        remove_btn.clicked.connect(self.remove_row)
        
        calculate_btn = QPushButton("🧮 Calculate")
        calculate_btn.clicked.connect(self.calculate_inventory)
        
        save_btn = QPushButton("💾 Save")
        save_btn.clicked.connect(self.save_data)
        
        load_btn = QPushButton("📥 Load")
        load_btn.clicked.connect(self.load_data)
        
        export_btn = QPushButton("📤 Export")
        export_btn.clicked.connect(self.export_data)
        
        layout.addWidget(add_btn)
        layout.addWidget(remove_btn)
        layout.addWidget(calculate_btn)
        layout.addStretch()
        layout.addWidget(save_btn)
        layout.addWidget(load_btn)
        layout.addWidget(export_btn)
        
        widget.setLayout(layout)
        return widget
        
    def load_sample_data(self):
        """بارگذاری داده نمونه"""
        if self.table.rowCount() == 0:
            sample_data = [
                ['Drill Pipe 5"', "Pipe", 100, 50, 30, 120, "joints", 50, 200],
                ['Casing 13 3/8"', "Casing", 200, 100, 50, 250, "joints", 100, 300],
                ["Cement", "Material", 500, 200, 300, 400, "sacks", 200, 600],
                ["Baryte", "Chemical", 1000, 500, 400, 1100, "lbs", 500, 1500],
                ["Bentonite", "Chemical", 800, 400, 200, 1000, "lbs", 400, 1200]
            ]
            
            for data in sample_data:
                self.add_row(data)
                
    def add_row(self, data=None):
        """اضافه کردن سطر"""
        if data is None:
            data = ["New Item", "Category", 0, 0, 0, 0, "pcs", 10, 100]
        
        self.table_manager.add_row(data)
        
    def remove_row(self):
        """حذف سطر"""
        self.table_manager.delete_row()
        
    def calculate_inventory(self):
        """محاسبه موجودی"""
        try:
            for row in range(self.table.rowCount()):
                try:
                    opening_item = self.table.item(row, 2)
                    received_item = self.table.item(row, 3)
                    used_item = self.table.item(row, 4)
                    
                    if opening_item and received_item and used_item:
                        opening = float(opening_item.text() or 0)
                        received = float(received_item.text() or 0)
                        used = float(used_item.text() or 0)
                        
                        remaining = opening + received - used
                        
                        remaining_item = self.table.item(row, 5)
                        if not remaining_item:
                            remaining_item = QTableWidgetItem()
                            self.table.setItem(row, 5, remaining_item)
                        
                        remaining_item.setText(f"{remaining:.2f}")
                        remaining_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                        
                        # رنگ‌بندی بر اساس سطح
                        min_item = self.table.item(row, 7)
                        max_item = self.table.item(row, 8)
                        
                        if min_item and max_item:
                            try:
                                min_level = float(min_item.text() or 0)
                                max_level = float(max_item.text() or 0)
                                
                                if remaining < min_level:
                                    remaining_item.setBackground(QColor(255, 200, 200))
                                elif remaining > max_level:
                                    remaining_item.setBackground(QColor(255, 255, 200))
                                else:
                                    remaining_item.setBackground(QColor(200, 255, 200))
                            except:
                                pass
                                
                except ValueError:
                    continue
                    
            QMessageBox.information(self, "Success", "Inventory calculated")
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Calculation failed: {str(e)}")
            
    def save_data(self):
        """ذخیره داده‌ها"""
        try:
            data = self.get_table_data()
            if self.parent_widget and hasattr(self.parent_widget, 'save_inventory'):
                return self.parent_widget.save_inventory(data)
                
            QMessageBox.information(self, "Success", "Inventory data saved")
            return True
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to save: {str(e)}")
            return False
            
    def load_data(self):
        """بارگذاری داده‌ها"""
        try:
            if self.parent_widget and hasattr(self.parent_widget, 'load_inventory'):
                data = self.parent_widget.load_inventory()
                if data:
                    self.load_table_data(data)
                    return True
                    
            QMessageBox.information(self, "Info", "No data to load")
            return False
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to load: {str(e)}")
            return False
            
    def export_data(self):
        """اکسپورت داده‌ها"""
        export_manager = ExportManager(self)
        result = export_manager.export_table_with_dialog(self.table, "inventory")
        if result:
            QMessageBox.information(self, "Success", f"Exported to {result}")
            
    def get_table_data(self):
        """دریافت داده‌های جدول"""
        data = []
        for row in range(self.table.rowCount()):
            row_data = []
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                row_data.append(item.text() if item else "")
            data.append(row_data)
        return data
        
    def load_table_data(self, data):
        """بارگذاری داده‌ها در جدول"""
        self.table.setRowCount(0)
        for row_data in data:
            row = self.table.rowCount()
            self.table.insertRow(row)
            for col, value in enumerate(row_data):
                if col < self.table.columnCount():
                    item = QTableWidgetItem(str(value))
                    self.table.setItem(row, col, item)


# ------------------------ Drill Pipe Tab ------------------------
class DrillPipeTab(QWidget):
    """تب مشخصات لوله حفاری"""
    
    def __init__(self, parent_widget=None, db_manager=None):
        super().__init__()
        self.parent_widget = parent_widget
        self.db_manager = db_manager
        self.table_manager = None
        self.init_ui()
        
    def init_ui(self):
        """راه‌اندازی رابط کاربری"""
        layout = QVBoxLayout()
        
        # توضیحات
        desc_label = QLabel("🔩 Drill Pipe Specifications - Detailed specs for all drill pipe in inventory")
        desc_label.setStyleSheet("font-size: 12px; color: #555; padding: 5px;")
        layout.addWidget(desc_label)
        
        # جدول لوله حفاری
        self.table = QTableWidget(0, 10)
        self.table.setHorizontalHeaderLabels([
            "Size & Weight", "Connection", "ID (in)", "Grade", 
            "TJ OD/ID", "Length (ft)", "Quantity", "Condition", 
            "Last Inspection", "Remarks"
        ])
        
        # Table Manager
        self.table_manager = TableManager(self.table, self)
        
        # دکمه‌ها
        button_widget = self.create_buttons()
        
        layout.addWidget(self.table)
        layout.addWidget(button_widget)
        
        self.setLayout(layout)
        
        # نمونه داده
        self.load_sample_data()
        
    def create_buttons(self):
        """ایجاد دکمه‌های تب"""
        widget = QWidget()
        layout = QHBoxLayout()
        
        add_btn = QPushButton("➕ Add Pipe")
        add_btn.clicked.connect(self.add_row)
        
        remove_btn = QPushButton("➖ Remove Pipe")
        remove_btn.clicked.connect(self.remove_row)
        
        save_btn = QPushButton("💾 Save")
        save_btn.clicked.connect(self.save_data)
        
        load_btn = QPushButton("📥 Load")
        load_btn.clicked.connect(self.load_data)
        
        export_btn = QPushButton("📤 Export")
        export_btn.clicked.connect(self.export_data)
        
        layout.addWidget(add_btn)
        layout.addWidget(remove_btn)
        layout.addStretch()
        layout.addWidget(save_btn)
        layout.addWidget(load_btn)
        layout.addWidget(export_btn)
        
        widget.setLayout(layout)
        return widget
        
    def load_sample_data(self):
        """بارگذاری داده نمونه"""
        if self.table.rowCount() == 0:
            sample_data = [
                ['5" 19.5#', "NC50", 4.276, "G-105", "6.5/4.0", 30, 200, "New", "2024-01-15", "Fresh from factory"],
                ['5" 19.5#', "NC50", 4.276, "G-105", "6.5/4.0", 30, 150, "Used - Good", "2024-02-01", "2nd string"],
                ['3 1/2" 13.3#', "NC38", 2.764, "G-105", "4.5/3.0", 30, 100, "New", "2024-01-20", "For small hole"],
                ['5" 19.5#', "FH", 4.276, "S-135", "6.5/4.0", 45, 50, "Premium", "2024-03-01", "Heavy weight"],
                ['6 5/8" 25.2#', "6 5/8 REG", 5.965, "G-105", "8.5/6.0", 30, 80, "Good", "2024-02-15", "For large hole"]
            ]
            
            for data in sample_data:
                self.add_row(data)
                
    def add_row(self, data=None):
        """اضافه کردن سطر"""
        if data is None:
            data = ['5" 19.5#', "NC50", "4.276", "G-105", 
                   "6.5/4.0", "30", "100", "New", 
                   datetime.now().strftime("%Y-%m-%d"), ""]
        
        self.table_manager.add_row(data)
        
    def remove_row(self):
        """حذف سطر"""
        self.table_manager.delete_row()
        
    def save_data(self):
        """ذخیره داده‌ها"""
        try:
            data = self.get_table_data()
            if self.parent_widget and hasattr(self.parent_widget, 'save_drill_pipe'):
                return self.parent_widget.save_drill_pipe(data)
                
            QMessageBox.information(self, "Success", "Drill pipe data saved")
            return True
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to save: {str(e)}")
            return False
            
    def load_data(self):
        """بارگذاری داده‌ها"""
        try:
            if self.parent_widget and hasattr(self.parent_widget, 'load_drill_pipe'):
                data = self.parent_widget.load_drill_pipe()
                if data:
                    self.load_table_data(data)
                    return True
                    
            QMessageBox.information(self, "Info", "No data to load")
            return False
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to load: {str(e)}")
            return False
            
    def export_data(self):
        """اکسپورت داده‌ها"""
        export_manager = ExportManager(self)
        result = export_manager.export_table_with_dialog(self.table, "drill_pipe")
        if result:
            QMessageBox.information(self, "Success", f"Exported to {result}")
            
    def get_table_data(self):
        """دریافت داده‌های جدول"""
        data = []
        for row in range(self.table.rowCount()):
            row_data = []
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                row_data.append(item.text() if item else "")
            data.append(row_data)
        return data
        
    def load_table_data(self, data):
        """بارگذاری داده‌ها در جدول"""
        self.table.setRowCount(0)
        for row_data in data:
            row = self.table.rowCount()
            self.table.insertRow(row)
            for col, value in enumerate(row_data):
                if col < self.table.columnCount():
                    item = QTableWidgetItem(str(value))
                    self.table.setItem(row, col, item)


# ------------------------ Solid Control Tab ------------------------
class SolidControlTab(QWidget):
    """تب کنترل جامدات"""
    
    def __init__(self, parent_widget=None, db_manager=None):
        super().__init__()
        self.parent_widget = parent_widget
        self.db_manager = db_manager
        self.table_manager = None
        self.init_ui()
        
    def init_ui(self):
        """راه‌اندازی رابط کاربری"""
        layout = QVBoxLayout()
        
        # توضیحات
        desc_label = QLabel("🔄 Solid Control Equipment - Monitor shakers, centrifuges, and other solid control equipment")
        desc_label.setStyleSheet("font-size: 12px; color: #555; padding: 5px;")
        layout.addWidget(desc_label)
        
        # جدول کنترل جامدات
        self.table = QTableWidget(0, 10)
        self.table.setHorizontalHeaderLabels([
            "Equipment", "Type", "Feed (bbl/hr)", "Hours Operated", 
            "Loss (bbl)", "Size/# Cones", "U.F (%)", "O.F (%)", 
            "Daily Hours", "Cumulative Hours"
        ])
        
        # Table Manager
        self.table_manager = TableManager(self.table, self)
        
        # دکمه‌ها
        button_widget = self.create_buttons()
        
        layout.addWidget(self.table)
        layout.addWidget(button_widget)
        
        self.setLayout(layout)
        
        # نمونه داده
        self.load_sample_data()
        
    def create_buttons(self):
        """ایجاد دکمه‌های تب"""
        widget = QWidget()
        layout = QHBoxLayout()
        
        add_btn = QPushButton("➕ Add Equipment")
        add_btn.clicked.connect(self.add_row)
        
        remove_btn = QPushButton("➖ Remove Equipment")
        remove_btn.clicked.connect(self.remove_row)
        
        save_btn = QPushButton("💾 Save")
        save_btn.clicked.connect(self.save_data)
        
        load_btn = QPushButton("📥 Load")
        load_btn.clicked.connect(self.load_data)
        
        export_btn = QPushButton("📤 Export")
        export_btn.clicked.connect(self.export_data)
        
        layout.addWidget(add_btn)
        layout.addWidget(remove_btn)
        layout.addStretch()
        layout.addWidget(save_btn)
        layout.addWidget(load_btn)
        layout.addWidget(export_btn)
        
        widget.setLayout(layout)
        return widget
        
    def load_sample_data(self):
        """بارگذاری داده نمونه"""
        if self.table.rowCount() == 0:
            sample_data = [
                ["Shaker 1", "Linear", 500, 24, 10, "4 Cones", 80, 20, 24, 1200],
                ["Shaker 2", "Linear", 450, 24, 8, "4 Cones", 85, 15, 24, 800],
                ["Centrifuge", "High Speed", 200, 12, 5, "2 Cones", 90, 10, 12, 500],
                ["Degasser", "Vacuum", 300, 24, 3, "N/A", 95, 5, 24, 600],
                ["Desilter", "12 Cone", 400, 18, 6, "12 Cones", 85, 15, 18, 400]
            ]
            
            for data in sample_data:
                self.add_row(data)
                
    def add_row(self, data=None):
        """اضافه کردن سطر"""
        if data is None:
            data = ["New Equipment", "Type", "500", "24", "10", 
                   "4 Cones", "80", "20", "24", "1200"]
        
        self.table_manager.add_row(data)
        
    def remove_row(self):
        """حذف سطر"""
        self.table_manager.delete_row()
        
    def save_data(self):
        """ذخیره داده‌ها"""
        try:
            data = self.get_table_data()
            if self.parent_widget and hasattr(self.parent_widget, 'save_solid_control'):
                return self.parent_widget.save_solid_control(data)
                
            QMessageBox.information(self, "Success", "Solid control data saved")
            return True
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to save: {str(e)}")
            return False
            
    def load_data(self):
        """بارگذاری داده‌ها"""
        try:
            if self.parent_widget and hasattr(self.parent_widget, 'load_solid_control'):
                data = self.parent_widget.load_solid_control()
                if data:
                    self.load_table_data(data)
                    return True
                    
            QMessageBox.information(self, "Info", "No data to load")
            return False
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to load: {str(e)}")
            return False
            
    def export_data(self):
        """اکسپورت داده‌ها"""
        export_manager = ExportManager(self)
        result = export_manager.export_table_with_dialog(self.table, "solid_control")
        if result:
            QMessageBox.information(self, "Success", f"Exported to {result}")
            
    def get_table_data(self):
        """دریافت داده‌های جدول"""
        data = []
        for row in range(self.table.rowCount()):
            row_data = []
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                row_data.append(item.text() if item else "")
            data.append(row_data)
        return data
        
    def load_table_data(self, data):
        """بارگذاری داده‌ها در جدول"""
        self.table.setRowCount(0)
        for row_data in data:
            row = self.table.rowCount()
            self.table.insertRow(row)
            for col, value in enumerate(row_data):
                if col < self.table.columnCount():
                    item = QTableWidgetItem(str(value))
                    self.table.setItem(row, col, item)


# ------------------------ Main Equipment Widget ------------------------
class EquipmentWidget(DrillWidgetBase):
    """ویجت اصلی تجهیزات که تب‌ها را مدیریت می‌کند"""
    
    def __init__(self, db_manager=None):
        super().__init__("EquipmentWidget", db_manager)
        self.db = db_manager
        self.current_well = None
        self.current_report = None
        self.equipment_data = {}
        self.tabs = {}
        self.init_ui()
        self.setup_shortcuts()
        
    def init_ui(self):
        """راه‌اندازی رابط کاربری"""
        main_layout = QVBoxLayout()
        
        # Header با انتخاب چاه و گزارش
        header_widget = self.create_header_widget()
        main_layout.addWidget(header_widget)
        
        # تب‌ها
        self.tab_widget = QTabWidget()
        self.create_tabs()
        main_layout.addWidget(self.tab_widget)
        
        # دکمه‌های پایین
        button_widget = self.create_button_widget()
        main_layout.addWidget(button_widget)
        
        self.setLayout(main_layout)
        
    def create_header_widget(self):
        """ایجاد هدر با انتخاب چاه و گزارش"""
        widget = QWidget()
        layout = QHBoxLayout()
        
        # انتخاب چاه
        well_label = QLabel("Well:")
        self.well_combo = QComboBox()
        self.well_combo.setMinimumWidth(200)
        self.well_combo.currentIndexChanged.connect(self.on_well_changed)
        
        # انتخاب گزارش
        report_label = QLabel("Report:")
        self.report_combo = QComboBox()
        self.report_combo.setMinimumWidth(200)
        self.report_combo.currentIndexChanged.connect(self.on_report_changed)
        
        # دکمه‌ها
        load_all_btn = QPushButton("📥 Load All")
        load_all_btn.clicked.connect(self.load_all_data)
        
        save_all_btn = QPushButton("💾 Save All")
        save_all_btn.clicked.connect(self.save_all_data)
        
        export_all_btn = QPushButton("📤 Export All")
        export_all_btn.clicked.connect(self.export_all_data)
        
        layout.addWidget(well_label)
        layout.addWidget(self.well_combo)
        layout.addWidget(report_label)
        layout.addWidget(self.report_combo)
        layout.addStretch()
        layout.addWidget(load_all_btn)
        layout.addWidget(save_all_btn)
        layout.addWidget(export_all_btn)
        
        widget.setLayout(layout)
        return widget
        
    def create_tabs(self):
        """ایجاد تب‌های مختلف"""
        # تب تجهیزات ریگ
        self.rig_tab = RigEquipmentTab(parent_widget=self, db_manager=self.db)
        self.tab_widget.addTab(self.rig_tab, "🏗️ Rig Equipment")
        self.tabs['rig'] = self.rig_tab
        
        # تب موجودی
        self.inventory_tab = InventoryTab(parent_widget=self, db_manager=self.db)
        self.tab_widget.addTab(self.inventory_tab, "📦 Inventory")
        self.tabs['inventory'] = self.inventory_tab
        
        # تب مشخصات لوله حفاری
        self.pipe_tab = DrillPipeTab(parent_widget=self, db_manager=self.db)
        self.tab_widget.addTab(self.pipe_tab, "🔩 Drill Pipe")
        self.tabs['pipe'] = self.pipe_tab
        
        # تب کنترل جامدات
        self.solid_tab = SolidControlTab(parent_widget=self, db_manager=self.db)
        self.tab_widget.addTab(self.solid_tab, "🔄 Solid Control")
        self.tabs['solid'] = self.solid_tab
        
    def create_button_widget(self):
        """ایجاد ویجت دکمه‌های پایین"""
        widget = QWidget()
        layout = QHBoxLayout()
        
        # دکمه رفرش
        refresh_btn = QPushButton("🔄 Refresh Data")
        refresh_btn.clicked.connect(self.refresh_data)
        
        # دکمه پاک کردن
        clear_btn = QPushButton("🗑️ Clear All")
        clear_btn.clicked.connect(self.clear_all_data)
        
        layout.addWidget(refresh_btn)
        layout.addStretch()
        layout.addWidget(clear_btn)
        
        widget.setLayout(layout)
        return widget
        
    # ==================== Event Handlers ====================
    def on_well_changed(self, index):
        """هنگام تغییر چاه"""
        if index >= 0:
            well_id = self.well_combo.itemData(index)
            self.current_well = well_id
            self.load_reports_for_well(well_id)
            
    def on_report_changed(self, index):
        """هنگام تغییر گزارش"""
        if index >= 0:
            report_id = self.report_combo.itemData(index)
            self.current_report = report_id
            
    def load_reports_for_well(self, well_id):
        """بارگذاری گزارش‌های چاه انتخاب شده"""
        self.report_combo.clear()
        if self.db and well_id:
            reports = self.db.get_daily_reports_by_well(well_id)
            for report in reports:
                self.report_combo.addItem(
                    f"Report {report['id']} - {report['report_date']}", 
                    report['id']
                )
                
    # ==================== Save/Load Methods ====================
    def save_rig_equipment(self, data):
        """ذخیره داده‌های تجهیزات ریگ"""
        try:
            self.equipment_data['rig_equipment'] = data
            if self.current_report and self.db:
                report_data = {'rig_equipment': data}
                self.db.update_daily_report_tab_data(
                    self.current_report, 'equipment', report_data
                )
            return True
        except Exception as e:
            logger.error(f"Error saving rig equipment: {e}")
            return False
            
    def load_rig_equipment(self):
        """بارگذاری داده‌های تجهیزات ریگ"""
        try:
            if self.current_report and self.db:
                report = self.db.get_daily_report_by_id(self.current_report)
                if report and report.get('equipment_data'):
                    return report['equipment_data'].get('rig_equipment')
            return None
        except Exception as e:
            logger.error(f"Error loading rig equipment: {e}")
            return None
            
    def save_inventory(self, data):
        """ذخیره داده‌های موجودی"""
        try:
            self.equipment_data['inventory'] = data
            if self.current_report and self.db:
                report_data = {'inventory': data}
                self.db.update_daily_report_tab_data(
                    self.current_report, 'equipment', report_data
                )
            return True
        except Exception as e:
            logger.error(f"Error saving inventory: {e}")
            return False
            
    def load_inventory(self):
        """بارگذاری داده‌های موجودی"""
        try:
            if self.current_report and self.db:
                report = self.db.get_daily_report_by_id(self.current_report)
                if report and report.get('equipment_data'):
                    return report['equipment_data'].get('inventory')
            return None
        except Exception as e:
            logger.error(f"Error loading inventory: {e}")
            return None
            
    def save_drill_pipe(self, data):
        """ذخیره داده‌های لوله حفاری"""
        try:
            self.equipment_data['drill_pipe'] = data
            if self.current_report and self.db:
                report_data = {'drill_pipe': data}
                self.db.update_daily_report_tab_data(
                    self.current_report, 'equipment', report_data
                )
            return True
        except Exception as e:
            logger.error(f"Error saving drill pipe: {e}")
            return False
            
    def load_drill_pipe(self):
        """بارگذاری داده‌های لوله حفاری"""
        try:
            if self.current_report and self.db:
                report = self.db.get_daily_report_by_id(self.current_report)
                if report and report.get('equipment_data'):
                    return report['equipment_data'].get('drill_pipe')
            return None
        except Exception as e:
            logger.error(f"Error loading drill pipe: {e}")
            return None
            
    def save_solid_control(self, data):
        """ذخیره داده‌های کنترل جامدات"""
        try:
            self.equipment_data['solid_control'] = data
            if self.current_report and self.db:
                report_data = {'solid_control': data}
                self.db.update_daily_report_tab_data(
                    self.current_report, 'equipment', report_data
                )
            return True
        except Exception as e:
            logger.error(f"Error saving solid control: {e}")
            return False
            
    def load_solid_control(self):
        """بارگذاری داده‌های کنترل جامدات"""
        try:
            if self.current_report and self.db:
                report = self.db.get_daily_report_by_id(self.current_report)
                if report and report.get('equipment_data'):
                    return report['equipment_data'].get('solid_control')
            return None
        except Exception as e:
            logger.error(f"Error loading solid control: {e}")
            return None
            
    # ==================== Main Operations ====================
    def save_all_data(self):
        """ذخیره همه داده‌ها"""
        try:
            # ذخیره هر تب
            for tab_name, tab in self.tabs.items():
                if hasattr(tab, 'save_data'):
                    tab.save_data()
                    
            self.status_manager.show_success(
                self.widget_name, 
                "All equipment data saved"
            )
            return True
            
        except Exception as e:
            self.status_manager.show_error(
                self.widget_name, 
                f"Failed to save all: {str(e)}"
            )
            return False
            
    def load_all_data(self):
        """بارگذاری همه داده‌ها"""
        try:
            if not self.current_report:
                QMessageBox.warning(self, "Warning", "Please select a report first")
                return False
                
            for tab_name, tab in self.tabs.items():
                if hasattr(tab, 'load_data'):
                    tab.load_data()
                    
            self.status_manager.show_success(
                self.widget_name, 
                "All equipment data loaded"
            )
            return True
            
        except Exception as e:
            self.status_manager.show_error(
                self.widget_name, 
                f"Failed to load all: {str(e)}"
            )
            return False
            
    def export_all_data(self):
        """اکسپورت همه داده‌ها"""
        try:
            from openpyxl import Workbook
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"equipment_export_all_{timestamp}.xlsx"
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Export All Equipment Data", 
                filename, "Excel Files (*.xlsx)"
            )
            
            if file_path:
                wb = Workbook()
                
                # اکسپورت هر تب
                for tab_name, tab in self.tabs.items():
                    if hasattr(tab, 'table'):
                        ws = wb.create_sheet(title=tab_name.capitalize())
                        
                        # نوشتن هدرها
                        table = tab.table
                        for col in range(table.columnCount()):
                            header = table.horizontalHeaderItem(col)
                            ws.cell(row=1, column=col+1, 
                                   value=header.text() if header else f"Column {col+1}")
                        
                        # نوشتن داده‌ها
                        for row in range(table.rowCount()):
                            for col in range(table.columnCount()):
                                item = table.item(row, col)
                                ws.cell(row=row+2, column=col+1, 
                                       value=item.text() if item else "")
                
                # حذف sheet پیش‌فرض
                if "Sheet" in wb.sheetnames:
                    std = wb["Sheet"]
                    wb.remove(std)
                
                wb.save(file_path)
                
                self.status_manager.show_success(
                    self.widget_name, 
                    f"All data exported to {file_path}"
                )
                
        except Exception as e:
            self.status_manager.show_error(
                self.widget_name, 
                f"Export failed: {str(e)}"
            )
            
    def refresh_data(self):
        """رفرش داده‌ها"""
        try:
            self.populate_wells()
            if self.current_well:
                self.load_reports_for_well(self.current_well)
                
            self.status_manager.show_message(
                self.widget_name, 
                "Data refreshed"
            )
            
        except Exception as e:
            self.status_manager.show_error(
                self.widget_name, 
                f"Refresh failed: {str(e)}"
            )
            
    def clear_all_data(self):
        """پاک کردن همه داده‌ها"""
        reply = QMessageBox.question(
            self, "Clear All", 
            "Are you sure you want to clear all tables?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            for tab_name, tab in self.tabs.items():
                if hasattr(tab, 'table'):
                    tab.table.setRowCount(0)
                    
            self.status_manager.show_message(
                self.widget_name, 
                "All tables cleared"
            )
            
    def populate_wells(self):
        """پر کردن لیست چاه‌ها"""
        if self.db:
            self.well_combo.clear()
            hierarchy = self.db.get_hierarchy()
            for company in hierarchy:
                for project in company['projects']:
                    for well in project['wells']:
                        self.well_combo.addItem(
                            f"{well['name']} ({project['name']})", 
                            well['id']
                        )
                        
    def setup_shortcuts(self):
        """تنظیم کلیدهای میانبر"""
        # Ctrl+S برای ذخیره همه
        save_shortcut = QShortcut(QKeySequence("Ctrl+S"), self)
        save_shortcut.activated.connect(self.save_all_data)
        
        # Ctrl+L برای بارگذاری همه
        load_shortcut = QShortcut(QKeySequence("Ctrl+L"), self)
        load_shortcut.activated.connect(self.load_all_data)
        
        # Ctrl+E برای اکسپورت
        export_shortcut = QShortcut(QKeySequence("Ctrl+E"), self)
        export_shortcut.activated.connect(self.export_all_data)
        
        # Ctrl+R برای رفرش
        refresh_shortcut = QShortcut(QKeySequence("Ctrl+R"), self)
        refresh_shortcut.activated.connect(self.refresh_data)
        
    def save_data(self):
        """ذخیره داده‌ها - برای استفاده توسط AutoSave"""
        return self.save_all_data()
        
    def load_data(self):
        """بارگذاری داده‌ها"""
        return self.load_all_data()
